import openpyxl
from PySide6.QtCore import QFileInfo
from collections import defaultdict

class ScoreLogic:
    def __init__(self):
        # 각 파일별로 path, headers, student_data, dirty, row_range를 저장
        self.files = []  # [{path, headers, student_data, dirty, row_range} ...]
        self.row_to_file_idx = []  # 테이블의 각 row가 어느 파일에 속하는지 인덱스 매핑
        self._cached_headers = None  # 헤더 캐싱
        self._cached_student_data = None  # 학생 데이터 캐싱
        self._cache_dirty = True  # 캐시 무효화 플래그

    def _invalidate_cache(self):
        """캐시를 무효화합니다."""
        self._cached_headers = None
        self._cached_student_data = None
        self._cache_dirty = True

    def load_excel_data(self, file_path):
        """
        엑셀 파일을 불러와서 self.files에 추가하고, row_to_file_idx를 갱신합니다.
        """
        if any(f['path'] == file_path for f in self.files):
            return False, "이미 추가된 파일입니다."

        try:
            # 메모리 효율적인 읽기
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            sheet = workbook.active

            # 헤더 최적화 - 한 번에 처리
            headers = []
            if sheet.max_row >= 2:
                rows_iter = sheet.iter_rows(min_row=1, max_row=2, values_only=True)
                row1 = next(rows_iter)
                row2 = next(rows_iter)
                headers = [f"{str(r1) if r1 else ''}\n{str(r2) if r2 else ''}".strip() 
                          for r1, r2 in zip(row1, row2)]
            elif sheet.max_row > 0:
                row1 = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [str(cell) if cell is not None else "" for cell in row1]

            # 학생 데이터 최적화 - 미리 할당된 리스트 사용
            student_data = []
            if sheet.max_row > 3:
                # 예상 행 수로 미리 할당
                expected_rows = sheet.max_row - 3
                student_data = [None] * expected_rows
                
                for idx, row in enumerate(sheet.iter_rows(min_row=4, values_only=True)):
                    if idx < expected_rows:
                        student_data[idx] = [str(val) if val is not None else "" for val in row]
                    else:
                        student_data.append([str(val) if val is not None else "" for val in row])
                
                # None 제거
                student_data = [row for row in student_data if row is not None]

            # row_range 계산 최적화
            start_row = sum(len(f['student_data']) for f in self.files)
            end_row = start_row + len(student_data) - 1 if student_data else start_row

            self.files.append({
                "path": file_path,
                "headers": headers,
                "student_data": student_data,
                "dirty": False,
                "row_range": (start_row, end_row)
            })

            self._update_row_to_file_idx_optimized()
            self._invalidate_cache()
            return True, "성공"
        except Exception as e:
            return False, f"엑셀 파일을 불러오는 중 오류가 발생했습니다:\n{e}"

    def _update_row_to_file_idx_optimized(self):
        """row_to_file_idx를 최적화하여 갱신합니다."""
        total_rows = sum(len(f['student_data']) for f in self.files)
        self.row_to_file_idx = [0] * total_rows
        
        current_row = 0
        for idx, f in enumerate(self.files):
            rows_count = len(f['student_data'])
            for i in range(rows_count):
                self.row_to_file_idx[current_row + i] = idx
            current_row += rows_count

    @property
    def headers(self):
        """헤더를 캐싱하여 반환합니다."""
        if self._cached_headers is None and self.files:
            self._cached_headers = self.files[0]['headers']
        return self._cached_headers or []

    @property
    def student_data(self):
        """학생 데이터를 캐싱하여 반환합니다."""
        if self._cached_student_data is None:
            # 제너레이터 대신 리스트 컴프리헨션으로 한 번에 생성
            all_data = []
            for f in self.files:
                all_data.extend(f['student_data'])
            self._cached_student_data = all_data
        return self._cached_student_data

    def update_score(self, row_idx, session_idx, score):
        """특정 테이블 row의 점수를 해당 파일의 데이터에 반영하고 dirty 표시"""
        if row_idx < 0 or row_idx >= len(self.row_to_file_idx):
            return
            
        file_idx = self.row_to_file_idx[row_idx]
        file = self.files[file_idx]
        file_row_idx = row_idx - file['row_range'][0]
        target_col = session_idx + 4
        
        if 0 <= file_row_idx < len(file['student_data']):
            student_row = file['student_data'][file_row_idx]
            
            # 필요한 경우에만 확장
            if target_col >= len(student_row):
                extend_size = target_col - len(student_row) + 1
                student_row.extend([""] * extend_size)
            
            # 점수 타입 변환 최적화
            if score != "":
                try:
                    f_value = float(score)
                    score = int(f_value) if f_value.is_integer() else f_value
                except (ValueError, TypeError):
                    pass
                    
            student_row[target_col] = score
            file['dirty'] = True
            
            # 캐시된 데이터도 업데이트
            if self._cached_student_data is not None:
                self._cached_student_data[row_idx][target_col] = score

    def save_to_excel(self):
        """
        dirty가 True인 파일만 저장합니다.
        """
        saved = 0
        errors = []
        
        for file in self.files:
            if not file['dirty']:
                continue
                
            try:
                # 메모리 효율적인 저장
                workbook = openpyxl.load_workbook(file['path'])
                sheet = workbook.active
                
                # 배치 업데이트를 위한 데이터 준비
                updates = []
                for r_idx, row_data in enumerate(file['student_data']):
                    for c_idx, cell_data in enumerate(row_data):
                        value = cell_data
                        # 숫자 변환 최적화
                        if value and value != '':
                            try:
                                f_value = float(value)
                                value = int(f_value) if f_value.is_integer() else f_value
                            except (ValueError, TypeError):
                                pass
                        updates.append((r_idx + 4, c_idx + 1, value))
                
                # 배치로 셀 업데이트
                for row_num, col_num, value in updates:
                    cell = sheet.cell(row=row_num, column=col_num)
                    cell.value = value
                    if isinstance(value, (int, float)):
                        cell.number_format = 'General'
                
                # 불필요한 행 삭제
                if sheet.max_row > len(file['student_data']) + 3:
                    sheet.delete_rows(len(file['student_data']) + 4, sheet.max_row)
                
                workbook.save(file['path'])
                workbook.close()  # 명시적으로 닫기
                file['dirty'] = False
                saved += 1
                
            except Exception as e:
                errors.append(f"{file['path']}: {e}")
        
        if errors:
            return False, f"일부 파일 저장 실패:\n" + "\n".join(errors)
        if saved == 0:
            return False, "저장할 변경사항이 없습니다."
        return True, f"{saved}개 파일에 변경 내용이 저장되었습니다."

    def clear_data(self):
        """데이터를 초기화합니다."""
        self.files.clear()
        self.row_to_file_idx.clear()
        self._invalidate_cache()