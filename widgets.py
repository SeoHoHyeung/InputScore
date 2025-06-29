import os
import sys
from PySide6.QtCore import Qt, Signal, QTimer
from PySide6.QtGui import QDragEnterEvent, QDropEvent
from PySide6.QtWidgets import (QLabel, QMessageBox, QWidget, QVBoxLayout, 
                             QTableWidget, QLineEdit, QPushButton, QComboBox, 
                             QHBoxLayout, QGroupBox, QHeaderView, QTableWidgetItem)
import openpyxl

class DropZone(QLabel):
    fileDropped = Signal(list)
    
    # 스타일 상수 - 클래스 변수로 메모리 효율성 향상
    STYLE_DEFAULT = ("QLabel { border: 2px dashed #aaa; border-radius: 5px; "
                    "background-color: #f9f9f9; color: #666; } "
                    "QLabel:hover { border-color: #0078d4; background-color: #f0f8ff; }")
    STYLE_DRAG_OVER = ("QLabel { border: 2px solid #0078d4; border-radius: 5px; "
                      "background-color: #e6f3ff; color: #666; }")
    
    # 지원되는 파일 확장자
    SUPPORTED_EXTENSIONS = (".xlsx", ".xls")

    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignCenter)
        self.setStyleSheet(self.STYLE_DEFAULT)
        
        # 드래그 상태 추적
        self._is_dragging = False
        
        # 스타일 변경 타이머 (debouncing)
        self._style_timer = QTimer()
        self._style_timer.setSingleShot(True)
        self._style_timer.timeout.connect(self._apply_default_style)

    def dragEnterEvent(self, event: QDragEnterEvent):
        """드래그 진입 이벤트 처리 최적화"""
        if event.mimeData().hasUrls():
            # 파일 확장자 미리 검증
            urls = event.mimeData().urls()
            has_valid_file = any(
                url.toLocalFile().lower().endswith(self.SUPPORTED_EXTENSIONS)
                for url in urls
            )
            
            if has_valid_file:
                event.acceptProposedAction()
                self._set_drag_style()
            else:
                event.ignore()
        else:
            event.ignore()

    def dragLeaveEvent(self, event):
        """드래그 이탈 이벤트 처리"""
        self._is_dragging = False
        # 지연된 스타일 복원 (깜빡임 방지)
        self._style_timer.start(50)

    def _set_drag_style(self):
        """드래그 스타일 설정"""
        if not self._is_dragging:
            self._is_dragging = True
            self._style_timer.stop()
            self.setStyleSheet(self.STYLE_DRAG_OVER)

    def _apply_default_style(self):
        """기본 스타일 적용"""
        if not self._is_dragging:
            self.setStyleSheet(self.STYLE_DEFAULT)

    def dropEvent(self, event: QDropEvent):
        """드롭 이벤트 처리 최적화"""
        self._is_dragging = False
        self.setStyleSheet(self.STYLE_DEFAULT)
        
        urls = event.mimeData().urls()
        if not urls:
            return
            
        # 파일 경로 필터링 최적화
        valid_files = []
        for url in urls:
            file_path = url.toLocalFile()
            if file_path and file_path.lower().endswith(self.SUPPORTED_EXTENSIONS):
                valid_files.append(file_path)
        
        if valid_files:
            self.fileDropped.emit(valid_files)
        else:
            QMessageBox.warning(self, "파일 형식 오류", "엑셀 파일(.xlsx, .xls)만 올려주세요.")


class MultiClassPanel(QWidget):
    """이동반 전용 패널 클래스 - 성능 최적화"""
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.files = []  # 업로드된 파일 정보 [{path, headers, data} ...]
        
        # 성능 최적화를 위한 변수들
        self._search_cache = {}  # 검색 결과 캐싱
        self._update_timer = QTimer()
        self._update_timer.setSingleShot(True)
        self._update_timer.timeout.connect(self._delayed_search)
        self._pending_search = None
        
        self.init_ui()

    def init_ui(self):
        """UI 초기화 최적화"""
        layout = QVBoxLayout(self)
        
        # 학생번호 입력란
        input_layout = QHBoxLayout()
        input_layout.addWidget(QLabel("학생번호 입력:"))
        
        self.student_number_input = QLineEdit()
        self.student_number_input.setPlaceholderText("예: 2")
        self.student_number_input.returnPressed.connect(self._schedule_search)
        input_layout.addWidget(self.student_number_input)
        
        layout.addLayout(input_layout)
        
        # 결과 테이블 최적화 설정
        self.student_table = QTableWidget()
        self.student_table.setColumnCount(2)
        self.student_table.setHorizontalHeaderLabels(["번호", "이름"])
        
        # 성능 최적화 설정
        header = self.student_table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        self.student_table.setAlternatingRowColors(True)
        self.student_table.setSortingEnabled(False)
        
        layout.addWidget(self.student_table)
        
        # 업로드된 파일명 표시
        self.file_list_label = QLabel("업로드된 파일: 없음")
        layout.addWidget(self.file_list_label)

    def _schedule_search(self):
        """검색 스케줄링 (debouncing)"""
        number = self.student_number_input.text().strip()
        self._pending_search = number
        self._update_timer.start(100)  # 100ms 지연

    def _delayed_search(self):
        """지연된 검색 실행"""
        if self._pending_search is not None:
            self.on_student_number_entered(self._pending_search)
            self._pending_search = None

    def on_file_dropped(self, file_path):
        """엑셀 파일 드롭 처리 최적화"""
        try:
            # 읽기 전용으로 최적화
            workbook = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
            sheet = workbook.active
            
            # 헤더 읽기
            headers = []
            if sheet.max_row > 0:
                first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
                headers = [str(cell) if cell is not None else "" for cell in first_row]
            
            # 데이터 읽기 최적화
            data = []
            if sheet.max_row > 1:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    row_data = [str(val) if val is not None else "" for val in row]
                    data.append(row_data)
            
            # 파일 정보 저장
            self.files.append({
                "path": file_path, 
                "headers": headers, 
                "data": data
            })
            
            # 캐시 무효화
            self._search_cache.clear()
            
            self.update_file_list_label()
            
        except Exception as e:
            QMessageBox.critical(self, "엑셀 로드 오류", 
                               f"엑셀 파일을 불러오는 중 오류가 발생했습니다:\n{e}")

    def update_file_list_label(self):
        """파일 리스트 라벨 업데이트"""
        if self.files:
            names = [os.path.basename(f["path"]) for f in self.files]
            display_text = "업로드된 파일: " + ", ".join(names)
            # 텍스트 길이 제한
            if len(display_text) > 50:
                display_text = display_text[:47] + "..."
            self.file_list_label.setText(display_text)
        else:
            self.file_list_label.setText("업로드된 파일: 없음")

    def on_student_number_entered(self, number=None):
        """학생번호 입력 처리 최적화"""
        if number is None:
            number = self.student_number_input.text().strip()
            
        if not number:
            self._clear_table()
            return
        
        # 캐시에서 확인
        if number in self._search_cache:
            results = self._search_cache[number]
        else:
            # 새로운 검색 수행
            results = self._search_student(number)
            
            # 캐시 크기 제한
            if len(self._search_cache) > 50:
                # 가장 오래된 항목들 제거
                old_keys = list(self._search_cache.keys())[:10]
                for key in old_keys:
                    del self._search_cache[key]
            
            self._search_cache[number] = results
        
        self._update_table(results, number)

    def _search_student(self, number):
        """학생 검색 최적화"""
        results = []
        for f in self.files:
            # B열(번호: index 1) 검색 최적화
            for row in f["data"]:
                if len(row) > 1 and row[1] == number:
                    # B열(번호: index 1), D열(이름: index 3) 추출
                    b_val = row[1] if len(row) > 1 else ""
                    d_val = row[3] if len(row) > 3 else ""
                    results.append([b_val, d_val])
        return results

    def _update_table(self, results, number):
        """테이블 업데이트 최적화"""
        table = self.student_table
        
        # 업데이트 차단
        table.setUpdatesEnabled(False)
        table.setRowCount(len(results))
        
        # 배치로 아이템 설정
        for i, row in enumerate(results):
            for j, text in enumerate(row):
                item = table.item(i, j)
                if not item:
                    item = QTableWidgetItem()
                    table.setItem(i, j, item)
                item.setText(text)
        
        table.setUpdatesEnabled(True)
        
        # 결과 없음 메시지
        if not results:
            QTimer.singleShot(0, lambda: QMessageBox.information(
                self, "검색 결과 없음", f"{number}번 학생을 찾을 수 없습니다."
            ))

    def _clear_table(self):
        """테이블 초기화"""
        table = self.student_table
        table.setUpdatesEnabled(False)
        table.clearContents()
        table.setRowCount(0)
        table.setUpdatesEnabled(True)

    def clear_data(self):
        """데이터 초기화"""
        self.files.clear()
        self._search_cache.clear()
        self._clear_table()
        self.update_file_list_label()

    def get_search_results_count(self):
        """검색 결과 수 반환"""
        return self.student_table.rowCount()

    def get_loaded_files_count(self):
        """로드된 파일 수 반환"""
        return len(self.files)